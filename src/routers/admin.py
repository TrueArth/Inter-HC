from fastapi import APIRouter, Depends, HTTPException, status
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime

from ..auth.auth import auth_handler
from ..dependencies import get_user_provider, get_interconsulta_provider, get_catalogo_provider
from ..controllers.user_controller import UserController
from ..controllers.catalogo_controller import CatalogoController

class SpecialtyCreate(BaseModel):
    nome: str

class SymptomCreate(BaseModel):
    nome: str
    pontuacao: int
    especialidade_id: int

class RuleCreate(BaseModel):
    sintoma_id: int
    especialidade_id: int
    pontuacao: int


router = APIRouter(prefix="/api", tags=["Admin"])

# --- Schemas ---

class AdminData(BaseModel):
    message: str
    user_groups: List[str]

class UserCreate(BaseModel):
    username: str = Field(..., description="Nome de usuário para login")
    password: str = Field(..., description="Senha do usuário")
    display_name: str = Field(..., description="Nome de exibição")
    role: str = Field(..., description="Papel: admin, medico ou regulador")
    email: Optional[str] = Field(None, description="Endereço de e-mail")

class UserUpdate(BaseModel):
    display_name: Optional[str] = Field(None, description="Nome de exibição")
    role: Optional[str] = Field(None, description="Papel: admin, medico ou regulador")
    email: Optional[str] = Field(None, description="Endereço de e-mail")

class UserResponse(BaseModel):
    id: int
    username: str
    display_name: str
    role: str
    email: Optional[str] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None

    class Config:
        from_attributes = True

# --- Dependências ---

async def verify_admin_group(current_user: dict = Depends(auth_handler.decode_token)):
    ADMIN_GROUP = "GLO-SEC-HCPE-SETISD"
    role = current_user.get("role")
    if role:
        if role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="Acesso negado: Operação restrita ao Administrador."
            )
    else:
        if ADMIN_GROUP not in current_user.get("groups", []):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="Acesso negado: Privilégios insuficientes."
            )
    return current_user

# --- Endpoints ---

@router.get("/admin-only-data", response_model=AdminData)
async def get_admin_data(current_user: dict = Depends(verify_admin_group)):
    """
    Retorna dados confidenciais apenas acessíveis por administradores.
    """
    return AdminData(
        message="This is highly confidential admin data!",
        user_groups=current_user.get("groups", [])
    )

@router.get("/admin/users", response_model=List[UserResponse])
async def get_users(
    provider = Depends(get_user_provider()),
    current_user = Depends(verify_admin_group)
):
    """
    Retorna a lista de usuários cadastrados no banco de dados local.
    """
    return await UserController.listar_usuarios(provider)

@router.post("/admin/users", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def create_user(
    payload: UserCreate,
    provider = Depends(get_user_provider()),
    current_user = Depends(verify_admin_group)
):
    """
    Cria um novo usuário local no banco de dados.
    """
    return await UserController.criar_usuario(payload.dict(), provider)

@router.put("/admin/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    payload: UserUpdate,
    provider = Depends(get_user_provider()),
    current_user = Depends(verify_admin_group)
):
    """
    Atualiza as informações de um usuário local.
    """
    return await UserController.atualizar_usuario(user_id, payload.dict(exclude_unset=True), provider)

@router.delete("/admin/users/{user_id}")
async def delete_user(
    user_id: int,
    provider = Depends(get_user_provider()),
    current_user = Depends(verify_admin_group)
):
    """
    Desativa um usuário local (Soft Delete).
    """
    return await UserController.cancelar_usuario(user_id, provider)


class SpecialtyStat(BaseModel):
    name: str
    count: int


class DoctorStat(BaseModel):
    name: str
    count: int


class AdminStatistics(BaseModel):
    top_specialty: SpecialtyStat
    specialties_distribution: List[SpecialtyStat]
    top_doctors: List[DoctorStat]
    inappropriate_doctors: List[DoctorStat]
    total_interconsultas: int
    tempo_medio_atendimento_segundos: float
    tempo_medio_atendimento_formatado: str
    especialidades_mais_pendencias: List[SpecialtyStat]


@router.get("/admin/statistics", response_model=AdminStatistics)
async def get_admin_statistics(
    interconsulta_provider = Depends(get_interconsulta_provider()),
    catalogo_provider = Depends(get_catalogo_provider()),
    user_provider = Depends(get_user_provider()),
    current_user = Depends(verify_admin_group)
):
    """
    Retorna estatísticas para o dashboard administrativo.
    """
    from collections import Counter
    
    # Busca todas as interconsultas ativas (onde deleted_at IS NULL)
    pedidos = await interconsulta_provider.listar_pedidos_ativos()
    
    # Dicionário de mapeamento de especialidades carregado do banco de dados
    try:
        especialidades = await catalogo_provider.listar_especialidades()
        especialidades_map = {esp["id"]: esp["nome"] for esp in especialidades}
        all_specialty_names = [esp["nome"] for esp in especialidades]
    except Exception:
        especialidades_map = {}
        all_specialty_names = []

    # Busca médicos cadastrados no banco de dados
    try:
        usuarios = await user_provider.listar_usuarios()
        all_doctors_names = [
            u.get("display_name") or u.get("username")
            for u in usuarios
            if u.get("role") == "medico" and u.get("deleted_at") is None
        ]
    except Exception:
        all_doctors_names = []
    
    specialties_counter = Counter({name: 0 for name in all_specialty_names})
    pending_specialties_counter = Counter({name: 0 for name in all_specialty_names})
    doctors_counter = Counter({name: 0 for name in all_doctors_names})
    indevidas_counter = Counter({name: 0 for name in all_doctors_names})
    
    for p in pedidos:
        esp_id = p.get("especialidade_id")
        esp_name = especialidades_map.get(esp_id, f"Especialidade {esp_id}")
        specialties_counter[esp_name] += 1
        
        status = p.get("status", "PENDENTE")
        if status == "PENDENTE":
            pending_specialties_counter[esp_name] += 1
            
        medico = p.get("medico_solicitante_crm") or "Desconhecido"
        doctors_counter[medico] += 1
        
        gravidade = p.get("gravidade", "VERDE").upper()
        if gravidade == "VERDE":
            indevidas_counter[medico] += 1
            
    # Formata retornos
    top_specialty_item = specialties_counter.most_common(1)
    top_specialty = {
        "name": top_specialty_item[0][0],
        "count": top_specialty_item[0][1]
    } if top_specialty_item else {"name": "Nenhuma", "count": 0}
    
    specialties_distribution = [
        {"name": k, "count": v} for k, v in specialties_counter.most_common()
    ]
    especialidades_mais_pendencias = [
        {"name": k, "count": v} for k, v in pending_specialties_counter.most_common()
    ]
    top_doctors = [
        {"name": k, "count": v} for k, v in doctors_counter.most_common()
    ]
    inappropriate_doctors = [
        {"name": k, "count": v} for k, v in indevidas_counter.most_common()
    ]
    
    # Calcula Tempo Médio de Atendimento da Marcação
    # Para pedidos com status "AGENDADO"
    agendados = [p for p in pedidos if p.get("status") == "AGENDADO"]
    delays = []
    for p in agendados:
        criado = p.get("criado_em")
        atualizado = p.get("atualizado_em")
        
        if isinstance(criado, str):
            try:
                criado = datetime.fromisoformat(criado.replace("Z", "+00:00"))
            except ValueError:
                criado = None
        if isinstance(atualizado, str):
            try:
                atualizado = datetime.fromisoformat(atualizado.replace("Z", "+00:00"))
            except ValueError:
                atualizado = None
                
        if criado and atualizado:
            if criado.tzinfo is not None and atualizado.tzinfo is None:
                atualizado = atualizado.replace(tzinfo=criado.tzinfo)
            elif criado.tzinfo is None and atualizado.tzinfo is not None:
                criado = criado.replace(tzinfo=atualizado.tzinfo)
                
            delta = (atualizado - criado).total_seconds()
            if delta >= 0:
                delays.append(delta)
                
    if delays:
        tempo_medio_segundos = sum(delays) / len(delays)
        horas = tempo_medio_segundos / 3600
        if horas < 1:
            tempo_medio_formatado = f"{int(tempo_medio_segundos / 60)} min"
        elif horas < 24:
            tempo_medio_formatado = f"{horas:.1f} horas"
        else:
            dias = int(horas // 24)
            restante_horas = horas % 24
            if restante_horas < 1:
                tempo_medio_formatado = f"{dias} dias"
            else:
                tempo_medio_formatado = f"{dias}d {restante_horas:.0f}h"
    else:
        tempo_medio_segundos = 0.0
        tempo_medio_formatado = "N/A"
        
    return {
        "top_specialty": top_specialty,
        "specialties_distribution": specialties_distribution,
        "top_doctors": top_doctors,
        "inappropriate_doctors": inappropriate_doctors,
        "total_interconsultas": len(pedidos),
        "tempo_medio_atendimento_segundos": tempo_medio_segundos,
        "tempo_medio_atendimento_formatado": tempo_medio_formatado,
        "especialidades_mais_pendencias": especialidades_mais_pendencias
    }


@router.get("/admin/statistics/export")
async def export_statistics_excel(
    interconsulta_provider = Depends(get_interconsulta_provider()),
    catalogo_provider = Depends(get_catalogo_provider()),
    user_provider = Depends(get_user_provider()),
    current_user = Depends(verify_admin_group)
):
    """
    Gera e faz streaming da planilha Excel (.xlsx) contendo os dados analíticos do portal.
    """
    import io
    import pandas as pd
    from fastapi.responses import StreamingResponse
    import logging
    from src.controllers.interconsulta_controller import resolver_nome_por_prep
    
    try:
        pedidos = await interconsulta_provider.listar_pedidos_ativos()
        
        try:
            especialidades = await catalogo_provider.listar_especialidades()
            especialidades_map = {esp["id"]: esp["nome"] for esp in especialidades}
            all_specialty_names = [esp["nome"] for esp in especialidades]
        except Exception:
            especialidades_map = {}
            all_specialty_names = []

        try:
            usuarios = await user_provider.listar_usuarios()
            all_doctors_names = [
                u.get("display_name") or u.get("username")
                for u in usuarios
                if u.get("role") == "medico" and u.get("deleted_at") is None
            ]
        except Exception:
            all_doctors_names = []
            
        # 1. Aba Geral (Resumo de KPIs)
        total_pedidos = len(pedidos)
        agendados = [p for p in pedidos if p.get("status") == "AGENDADO"]
        pendentes = [p for p in pedidos if p.get("status") == "PENDENTE"]
        verdes = [p for p in pedidos if p.get("gravidade", "VERDE").upper() == "VERDE"]
        
        # Specialties, Doctors, Inappropriates counters
        from collections import Counter
        specialties_counter = Counter({name: 0 for name in all_specialty_names})
        pending_counter = Counter({name: 0 for name in all_specialty_names})
        doctors_counter = Counter({name: 0 for name in all_doctors_names})
        indevidas_counter = Counter({name: 0 for name in all_doctors_names})
        
        for p in pedidos:
            esp_id = p.get("especialidade_id")
            esp_name = especialidades_map.get(esp_id, f"Especialidade {esp_id}")
            specialties_counter[esp_name] += 1
            if p.get("status") == "PENDENTE":
                pending_counter[esp_name] += 1
                
            medico = p.get("medico_solicitante_crm") or "Desconhecido"
            doctors_counter[medico] += 1
            
            gravidade = p.get("gravidade", "VERDE").upper()
            if gravidade == "VERDE":
                indevidas_counter[medico] += 1

        top_specialty_item = specialties_counter.most_common(1)
        top_specialty_name = top_specialty_item[0][0] if top_specialty_item else "Nenhuma"
        top_specialty_count = top_specialty_item[0][1] if top_specialty_item else 0
        
        delays = []
        for p in agendados:
            criado = p.get("criado_em")
            atualizado = p.get("atualizado_em")
            if isinstance(criado, str):
                try: criado = datetime.fromisoformat(criado.replace("Z", "+00:00"))
                except ValueError: criado = None
            if isinstance(atualizado, str):
                try: atualizado = datetime.fromisoformat(atualizado.replace("Z", "+00:00"))
                except ValueError: atualizado = None
            if criado and atualizado:
                if criado.tzinfo is not None and atualizado.tzinfo is None:
                    atualizado = atualizado.replace(tzinfo=criado.tzinfo)
                elif criado.tzinfo is None and atualizado.tzinfo is not None:
                    criado = criado.replace(tzinfo=atualizado.tzinfo)
                delta = (atualizado - criado).total_seconds()
                if delta >= 0:
                    delays.append(delta)
                    
        tma_str = "N/A"
        if delays:
            avg_sec = sum(delays) / len(delays)
            horas = avg_sec / 3600
            if horas < 1:
                tma_str = f"{int(avg_sec / 60)} min"
            elif horas < 24:
                tma_str = f"{horas:.1f} horas"
            else:
                dias = int(horas // 24)
                restante_horas = horas % 24
                if restante_horas < 1:
                    tma_str = f"{dias} dias"
                else:
                    tma_str = f"{dias}d {restante_horas:.0f}h"
                
        df_geral = pd.DataFrame([
            {"Indicador": "Total de Interconsultas", "Valor": total_pedidos, "Descrição / Insight": "Volume acumulado de solicitações ativas no portal"},
            {"Indicador": "Tempo Médio de Atendimento da Marcação", "Valor": tma_str, "Descrição / Insight": "Média de tempo entre abertura do pedido e marcação de consulta"},
            {"Indicador": "Especialidade Mais Solicitada (Nome)", "Valor": top_specialty_name, "Descrição / Insight": "Especialidade com maior volume de encaminhamentos"},
            {"Indicador": "Especialidade Mais Solicitada (Volume)", "Valor": top_specialty_count, "Descrição / Insight": "Total de pedidos da especialidade mais demandada"},
            {"Indicador": "Total de Casos Indevidos (Verde)", "Valor": len(verdes), "Descrição / Insight": "Encaminhamentos de baixa complexidade que poderiam ser resolvidos na APS"},
            {"Indicador": "Solicitações Pendentes", "Valor": len(pendentes), "Descrição / Insight": "Pacientes aguardando triagem/marcação na fila reguladora"},
            {"Indicador": "Solicitações Agendadas", "Valor": len(agendados), "Descrição / Insight": "Consultas agendadas com sucesso"}
        ])
        
        # 2. Aba de Demandas por Especialidade
        esp_rows = []
        for k, v in specialties_counter.most_common():
            part = (v / total_pedidos * 100) if total_pedidos > 0 else 0.0
            if v >= 10:
                classif = "Muito Alta"
            elif v >= 5:
                classif = "Alta"
            elif v > 0:
                classif = "Média"
            else:
                classif = "Nenhuma"
            esp_rows.append({
                "Especialidade": k,
                "Total de Solicitações": v,
                "Participação (%)": f"{part:.1f}%",
                "Classificação de Demanda": classif
            })
        df_esp = pd.DataFrame(esp_rows)
        if df_esp.empty:
            df_esp = pd.DataFrame(columns=["Especialidade", "Total de Solicitações", "Participação (%)", "Classificação de Demanda"])
            
        # 3. Aba de Pendências por Especialidade
        pend_rows = []
        for k, v in pending_counter.most_common():
            if v >= 5:
                alerta = "Crítico"
                acao = "Alocar médicos reguladores imediatamente"
            elif v > 0:
                alerta = "Médio"
                acao = "Monitorar fila de regulação"
            else:
                alerta = "Controlado"
                acao = "Nenhuma ação necessária"
            pend_rows.append({
                "Especialidade": k,
                "Solicitações Pendentes": v,
                "Nível de Alerta": alerta,
                "Ação Recomendada": acao
            })
        df_pendencias = pd.DataFrame(pend_rows)
        if df_pendencias.empty:
            df_pendencias = pd.DataFrame(columns=["Especialidade", "Solicitações Pendentes", "Nível de Alerta", "Ação Recomendada"])
            
        # 4. Aba de Médicos Mais Solicitantes
        doc_rows = []
        for k, v in doctors_counter.most_common():
            if v >= 10:
                freq = "Alta Frequência"
            elif v >= 3:
                freq = "Moderada"
            elif v > 0:
                freq = "Baixa"
            else:
                freq = "Sem Solicitações"
            doc_rows.append({
                "Médico Solicitante": k,
                "Volume de Solicitações": v,
                "Frequência de Uso": freq
            })
        df_doctors = pd.DataFrame(doc_rows)
        if df_doctors.empty:
            df_doctors = pd.DataFrame(columns=["Médico Solicitante", "Volume de Solicitações", "Frequência de Uso"])
            
        # 5. Aba de Casos Indevidos por Médico
        inappropriate_rows = []
        for k, v in indevidas_counter.most_common():
            total_doc = doctors_counter.get(k, 0)
            percentage = (v / total_doc * 100) if total_doc > 0 else 0.0
            if percentage >= 50.0 and total_doc >= 3:
                grau = "Crítico (Solicitar Reciclagem)"
            elif percentage > 0.0:
                grau = "Atenção (Rever Encaminhamentos)"
            else:
                grau = "Adequado"
            inappropriate_rows.append({
                "Médico Solicitante": k,
                "Casos Indevidos (Verde)": v,
                "Total de Solicitações": total_doc,
                "Índice de Inadequação (%)": f"{percentage:.1f}%",
                "Grau de Atenção": grau
            })
        df_inappropriate = pd.DataFrame(inappropriate_rows)
        if df_inappropriate.empty:
            df_inappropriate = pd.DataFrame(columns=["Médico Solicitante", "Casos Indevidos (Verde)", "Total de Solicitações", "Índice de Inadequação (%)", "Grau de Atenção"])
            
        # 6. Aba de Solicitações Detalhadas
        rows_detalhe = []
        for p in pedidos:
            esp_id = p.get("especialidade_id")
            esp_name = especialidades_map.get(esp_id, f"Especialidade {esp_id}")
            
            criado_em_str = ""
            if p.get("criado_em"):
                c_dt = p.get("criado_em")
                if isinstance(c_dt, str):
                    criado_em_str = c_dt
                else:
                    criado_em_str = c_dt.strftime("%Y-%m-%d %H:%M:%S")
                    
            prep_decrypted = p.get("paciente_prep", "")
            paciente_nome = resolver_nome_por_prep(prep_decrypted)
            
            rows_detalhe.append({
                "ID Solicitação": p.get("id"),
                "PREP Paciente": prep_decrypted,
                "Nome Paciente": paciente_nome,
                "Contato Paciente": p.get("paciente_contato") or "Não Informado",
                "Médico Solicitante": p.get("medico_solicitante_crm", ""),
                "Especialidade": esp_name,
                "Gravidade": p.get("gravidade", ""),
                "Status": p.get("status", ""),
                "Marcado Por": p.get("marcado_por") or "N/A",
                "Data de Criação": criado_em_str
            })
            
        df_detalhes = pd.DataFrame(rows_detalhe)
        if df_detalhes.empty:
            df_detalhes = pd.DataFrame(columns=["ID Solicitação", "PREP Paciente", "Nome Paciente", "Contato Paciente", "Médico Solicitante", "Especialidade", "Gravidade", "Status", "Marcado Por", "Data de Criação"])
            
        # 7. Aba de Usuários Cadastrados
        user_rows = []
        for u in usuarios:
            status_usuario = "Desativado" if u.get("deleted_at") else "Ativo"
            user_rows.append({
                "ID Usuário": u.get("id"),
                "Nome de Usuário (login)": u.get("username"),
                "Nome Exibido": u.get("display_name"),
                "Email": u.get("email"),
                "Função / Perfil (Role)": u.get("role"),
                "Status": status_usuario
            })
        df_users = pd.DataFrame(user_rows)
        if df_users.empty:
            df_users = pd.DataFrame(columns=["ID Usuário", "Nome de Usuário (login)", "Nome Exibido", "Email", "Função / Perfil (Role)", "Status"])
            
        # 8. Aba de Catálogo de Especialidades
        specialties_catalog_rows = []
        for esp in especialidades:
            specialties_catalog_rows.append({
                "ID Especialidade": esp.get("id"),
                "Nome da Especialidade": esp.get("nome")
            })
        df_specialties_catalog = pd.DataFrame(specialties_catalog_rows)
        if df_specialties_catalog.empty:
            df_specialties_catalog = pd.DataFrame(columns=["ID Especialidade", "Nome da Especialidade"])
            
        # 9. Aba de Catálogo de Sintomas
        try:
            sintomas = await catalogo_provider.listar_sintomas()
        except Exception:
            sintomas = []
            
        symptoms_catalog_rows = []
        for s in sintomas:
            esp_id = s.get("especialidade_id")
            esp_name = especialidades_map.get(esp_id, f"Especialidade {esp_id}")
            symptoms_catalog_rows.append({
                "ID Sintoma": s.get("id"),
                "Nome do Sintoma": s.get("nome"),
                "Pontuação Padrão (Score)": s.get("pontuacao"),
                "Especialidade Vinculada": esp_name
            })
        df_symptoms_catalog = pd.DataFrame(symptoms_catalog_rows)
        if df_symptoms_catalog.empty:
            df_symptoms_catalog = pd.DataFrame(columns=["ID Sintoma", "Nome do Sintoma", "Pontuação Padrão (Score)", "Especialidade Vinculada"])
            
        # Gerar arquivo Excel na memória com formatação openpyxl
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_geral.to_excel(writer, sheet_name='Indicadores Gerais', index=False)
            df_detalhes.to_excel(writer, sheet_name='Fila Reguladora', index=False)
            df_esp.to_excel(writer, sheet_name='Volume por Especialidade', index=False)
            df_pendencias.to_excel(writer, sheet_name='Pendencias por Especialidade', index=False)
            df_doctors.to_excel(writer, sheet_name='Medicos Mais Solicitantes', index=False)
            df_inappropriate.to_excel(writer, sheet_name='Casos Indevidos por Medico', index=False)
            df_users.to_excel(writer, sheet_name='Controle de Usuarios', index=False)
            df_specialties_catalog.to_excel(writer, sheet_name='Catalogo Especialidades', index=False)
            df_symptoms_catalog.to_excel(writer, sheet_name='Catalogo Sintomas', index=False)
            
            # Formatação visual das planilhas
            workbook = writer.book
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            
            alert_red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            alert_amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            alert_green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
            
            for name in workbook.sheetnames:
                ws = workbook[name]
                ws.views.sheetView[0].showGridLines = True
                
                # Ajusta colunas
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
                    
                # Formata Header
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                # Bordas nas células de dados
                for row in range(2, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                        
                # Formatação condicional baseada nos dados objetivos (não-numéricos)
                if name == "Casos Indevidos por Medico":
                    for row in range(2, ws.max_row + 1):
                        cell_grau = ws.cell(row=row, column=5)
                        if cell_grau.value:
                            if "Crítico" in str(cell_grau.value):
                                cell_grau.fill = alert_red_fill
                            elif "Atenção" in str(cell_grau.value):
                                cell_grau.fill = alert_amber_fill
                            elif "Adequado" in str(cell_grau.value):
                                cell_grau.fill = alert_green_fill
                                
                elif name == "Pendencias por Especialidade":
                    for row in range(2, ws.max_row + 1):
                        cell_alerta = ws.cell(row=row, column=3)
                        if cell_alerta.value:
                            if "Crítico" in str(cell_alerta.value):
                                cell_alerta.fill = alert_red_fill
                            elif "Médio" in str(cell_alerta.value):
                                cell_alerta.fill = alert_amber_fill
                            elif "Controlado" in str(cell_alerta.value):
                                cell_alerta.fill = alert_green_fill
                                
                elif name == "Volume por Especialidade":
                    for row in range(2, ws.max_row + 1):
                        cell_class = ws.cell(row=row, column=4)
                        if cell_class.value:
                            if "Muito Alta" in str(cell_class.value) or "Alta" in str(cell_class.value):
                                cell_class.fill = alert_amber_fill
                            elif "Média" in str(cell_class.value):
                                cell_class.fill = alert_green_fill
                                
        output.seek(0)
        
        username = current_user.get("username") or current_user.get("name") or "admin"
        logger = logging.getLogger("audit")
        logger.warning(
            f"AUDITORIA: Usuario '{username}' exportou os dados analiticos para planilha Excel (analytics_interhc.xlsx)."
        )
        
        headers = {
            'Content-Disposition': 'attachment; filename="analytics_interhc.xlsx"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao exportar dados para Excel: {str(e)}"
        )


# --- Dynamic Catalog CRUD Management Endpoints ---

@router.get("/admin/especialidades")
async def get_especialidades(
    provider = Depends(get_catalogo_provider()),
    current_user = Depends(verify_admin_group)
):
    """Retorna todas as especialidades ativas."""
    return await CatalogoController.listar_especialidades(provider)

@router.post("/admin/especialidades")
async def create_especialidade(
    payload: SpecialtyCreate,
    provider = Depends(get_catalogo_provider()),
    current_user = Depends(verify_admin_group)
):
    """Cadastra uma nova especialidade."""
    return await CatalogoController.criar_especialidade(payload.nome, provider)

@router.delete("/admin/especialidades/{especialidade_id}")
async def delete_especialidade(
    especialidade_id: int,
    provider = Depends(get_catalogo_provider()),
    current_user = Depends(verify_admin_group)
):
    """Inativa uma especialidade (Soft Delete)."""
    return await CatalogoController.inativar_especialidade(especialidade_id, provider)


@router.get("/admin/sintomas")
async def get_sintomas(
    provider = Depends(get_catalogo_provider()),
    current_user = Depends(verify_admin_group)
):
    """Retorna todos os sintomas ativos."""
    return await CatalogoController.listar_sintomas(provider)

@router.post("/admin/sintomas")
async def create_sintoma(
    payload: SymptomCreate,
    provider = Depends(get_catalogo_provider()),
    current_user = Depends(verify_admin_group)
):
    """Cadastra um novo sintoma e o associa a uma especialidade."""
    return await CatalogoController.criar_sintoma(payload.nome, payload.pontuacao, payload.especialidade_id, provider)

@router.delete("/admin/sintomas/{sintoma_id}")
async def delete_sintoma(
    sintoma_id: int,
    provider = Depends(get_catalogo_provider()),
    current_user = Depends(verify_admin_group)
):
    """Inativa um sintoma (Soft Delete)."""
    return await CatalogoController.inativar_sintoma(sintoma_id, provider)


@router.get("/admin/regras")
async def get_regras(
    provider = Depends(get_catalogo_provider()),
    current_user = Depends(verify_admin_group)
):
    """Retorna todas as regras de gravidade ativas."""
    return await CatalogoController.listar_regras_gravidade(provider)

@router.post("/admin/regras")
async def create_regra(
    payload: RuleCreate,
    provider = Depends(get_catalogo_provider()),
    current_user = Depends(verify_admin_group)
):
    """Cria ou atualiza uma regra de pontuação override."""
    return await CatalogoController.criar_regra_gravidade(
        payload.sintoma_id, 
        payload.especialidade_id, 
        payload.pontuacao, 
        provider
    )

@router.delete("/admin/regras/{sintoma_id}/{especialidade_id}")
async def delete_regra(
    sintoma_id: int,
    especialidade_id: int,
    provider = Depends(get_catalogo_provider()),
    current_user = Depends(verify_admin_group)
):
    """Inativa/remove uma regra de gravidade override."""
    return await CatalogoController.inativar_regra_gravidade(sintoma_id, especialidade_id, provider)

